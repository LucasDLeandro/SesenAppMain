import datetime
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from io import BytesIO

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.db.models import Count, Q
from django.template.loader import get_template
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from xhtml2pdf import pisa

from .models import (
    TelefoneSolicitacao, AparelhoVoip, RemessaManutencao, CriarSenha, 
    AparelhoManutencao, ContratoColaborador, PadraoSenhaTelefonia, 
    PadraoTutorialTelefonia, PadraoEmailTelefonia, EmprestimoEvento,
    TelefoneSolicitacaoAnexo
)
from .serializers import (
    TelefoneSolicitacaoSerializer,
    RemessaManutencaoSerializer, CriarSenhaSerializer, 
    ContratoColaboradorSerializer, EmprestimoEventoSerializer
)

from notificacoes.services import auto_message
from notificacoes.models.template_notificacao import TemplateMessage
from django.contrib.auth.models import User

# VIEWS TRADICIONAIS (TEMPLATES)
def main_telefonia(request):
    """Renderiza a tela principal de telefonia."""
    aparelhos_disponiveis = AparelhoVoip.objects.filter(status='estoque', integridade='funciona').count()
    aparelhos_instalados = AparelhoVoip.objects.filter(status='instalado').count()
    aparelhos_defeituosos = AparelhoVoip.objects.filter(status='defeituoso').count()
    aparelhos_manutencao = AparelhoVoip.objects.filter(status='manutencao').count()
    
    solicitacoes_pendentes = TelefoneSolicitacao.objects.exclude(status='concluida').count()
    solicitacoes_concluidas = TelefoneSolicitacao.objects.filter(status='concluida').count()
    
    eventos_ativos = EmprestimoEvento.objects.filter(status='em_andamento').count()
    
    remessas_total = RemessaManutencao.objects.count()
    total_senhas = CriarSenha.objects.count()
    
    tecnicos = User.objects.filter(groups__name__icontains='Telefonia')

    context = {
        'aparelhos_disponiveis': aparelhos_disponiveis,
        'aparelhos_instalados': aparelhos_instalados,
        'aparelhos_defeituosos': aparelhos_defeituosos,
        'solicitacoes_pendentes': solicitacoes_pendentes,
        'solicitacoes_concluidas': solicitacoes_concluidas,
        'eventos_ativos': eventos_ativos,
        'remessas_total': remessas_total,
        'total_senhas': total_senhas,
        'tecnicos': tecnicos,
    }
    return render(request, 'telefonia/main_telefonia.html', context)


# DASHBOARD STATS
def dashboard_stats(request):
    """Retorna dados estatísticos para os gráficos da telefonia"""
    solicitacoes_status = TelefoneSolicitacao.objects.values('status').annotate(total=Count('id'))
    aparelhos_status = AparelhoVoip.objects.values('status').annotate(total=Count('id'))
    senhas_categoria = CriarSenha.objects.values('categoria').annotate(total=Count('id'))
    
    return JsonResponse({
        'solicitacoes': list(solicitacoes_status),
        'aparelhos': list(aparelhos_status),
        'senhas': list(senhas_categoria)
    })


# DRF VIEWSETS
class TelefoneSolicitacaoViewSet(viewsets.ModelViewSet):
    queryset = TelefoneSolicitacao.objects.all().order_by('-data')
    serializer_class = TelefoneSolicitacaoSerializer

    def perform_create(self, serializer):
        solicitacao = serializer.save()
        template = TemplateMessage.objects.filter(tipo_evento='tel_solicitacao_aparelho', is_ativo=True).first()
        if template:
            from notificacoes.models.contato_notificacao import Contato
            from notificacoes.services import disparar_notificacao_contato
            
            contatos = Contato.objects.filter(is_ativo=True, notifica_telefonia=True)
            enviados = set()
            for contato in contatos:
                # Verificação de duplicidade por telefone ou e-mail
                chave_duplicidade = getattr(contato, '_telefone_sanitizado', contato.telefone) or (contato.pessoa.email if contato.pessoa else None)
                if chave_duplicidade:
                    if chave_duplicidade in enviados:
                        continue
                    enviados.add(chave_duplicidade)
                
                texto = template.base_text
                try:
                    text = texto.format(
                        tecnico=contato.nome,
                        protocolo=solicitacao.protocolo or 'N/A',
                        unidade=solicitacao.unidade or 'N/A',
                        sigla_unidade=solicitacao.sigla_unidade or 'N/A',
                        ramal=solicitacao.ramal or 'N/A',
                        local=solicitacao.local or 'N/A',
                        qnt_solicitada=solicitacao.qnt_solicitada or 0,
                        solicitante=solicitacao.solicitante or 'N/A',
                    )
                    assunto = f"Nova Solicitação de Aparelho - {solicitacao.protocolo or 'N/A'}"
                    disparar_notificacao_contato(contato, text, text, assunto)
                except Exception as e:
                    print(f"Erro ao formatar/enviar mensagem: {e}")

    def perform_update(self, serializer):
        solicitacao = serializer.save()
        
        has_new_files = False
        import re
        for key, file in self.request.FILES.items():
            match = re.match(r'^pdf_termos_edit_(\d+)$', key)
            if match:
                has_new_files = True
                index = int(match.group(1))
                
                anexo, created = TelefoneSolicitacaoAnexo.objects.get_or_create(
                    solicitacao=solicitacao,
                    ordem=index
                )
                
                if not created and anexo.arquivo:
                    anexo.arquivo.delete(save=False)
                    
                safe_termo = ''.join(c for c in (solicitacao.termo_transferencia_interna or 'termo') if c.isalnum() or c in (' ', '-', '_')).strip().replace(' ', '_')
                file.name = f"{safe_termo}_{solicitacao.protocolo}_{index}.pdf"
                anexo.arquivo = file
                anexo.save()
                
        if has_new_files:
            if hasattr(solicitacao, 'midia') and solicitacao.midia:
                solicitacao.midia.delete(save=False)
                solicitacao.midia = None
            if hasattr(solicitacao, 'pdf_termo') and solicitacao.pdf_termo:
                solicitacao.pdf_termo.delete(save=False)
                solicitacao.pdf_termo = None
            solicitacao.save()

    @action(detail=True, methods=['patch'])
    def concluir(self, request, pk=None):
        solicitacao = self.get_object()
        
        # Pega os outros campos
        solicitacao.tecnico_responsavel = request.data.get('tecnico_responsavel', solicitacao.tecnico_responsavel)
        solicitacao.relatorio = request.data.get('relatorio', solicitacao.relatorio)
        
        if 'data_instalacao' in request.data:
            data_inst_str = request.data.get('data_instalacao')
            if data_inst_str:
                dt = parse_datetime(data_inst_str)
                if dt:
                    if timezone.is_naive(dt):
                        dt = timezone.make_aware(dt)
                    solicitacao.data_instalacao = dt
            else:
                solicitacao.data_instalacao = None
                
        solicitacao.termo_transferencia_interna = request.data.get('termo_transferencia_interna', solicitacao.termo_transferencia_interna)
        
        instalacoes = request.data.get('instalacoes', [])
        
        ramais_agrupados = []
        locais_agrupados = []
        aparelhos_salvar = []
        
        if instalacoes:
            from telefonia.models import AparelhoVoip
            for inst in instalacoes:
                ap_id = inst.get('aparelho_id')
                ramal = inst.get('ramal')
                local = inst.get('local')
                mac_address = inst.get('mac_address')
                
                if ap_id:
                    try:
                        ap = AparelhoVoip.objects.get(id=ap_id)
                        ap.status = 'instalado'
                        ap.ramal = ramal or ap.ramal
                        ap.sala = local or ap.sala
                        if mac_address:
                            ap.mac_address = mac_address
                        ap.save()
                        
                        aparelhos_salvar.append(ap)
                        
                        if ap.ramal:
                            ramais_agrupados.append(ap.ramal)
                        if ap.sala:
                            locais_agrupados.append(ap.sala)
                    except AparelhoVoip.DoesNotExist:
                        pass
                        
            solicitacao.aparelhos.set(aparelhos_salvar)
            
        solicitacao.status = 'aguardando_supervisor_aparelho'
        if ramais_agrupados:
            solicitacao.ramal = ", ".join(ramais_agrupados)
        if locais_agrupados:
            solicitacao.local = ", ".join(locais_agrupados)
            
        solicitacao.save()
        return Response({'status': 'Solicitação técnica concluída com sucesso. Aguardando fase administrativa.'}, status=status.HTTP_200_OK)

    @transaction.atomic
    @action(detail=True, methods=['post'])
    def finalizar_administrativo(self, request, pk=None):
        solicitacao = self.get_object()
        
        if solicitacao.status != 'aguardando_supervisor_aparelho':
            return Response({'error': 'A solicitação não está na fase administrativa.'}, status=status.HTTP_400_BAD_REQUEST)
        termo = request.data.get('termo_transferencia_interna')
        pdf_termos = request.FILES.getlist('pdf_termos')
        
        if not termo or str(termo).strip() == '':
            return Response({'error': 'O preenchimento do número do Termo de Transferência Interna é obrigatório.'}, status=status.HTTP_400_BAD_REQUEST)
        
        solicitacao.termo_transferencia_interna = termo
        solicitacao.status = 'concluida'
        solicitacao.save()
        
        import os
        import re
        safe_termo = re.sub(r'[^a-zA-Z0-9_-]', '', str(termo).replace('/', '').replace('\\', '').replace(' ', ''))
        
        # Limpar anexos antigos caso seja uma resubmissão para evitar duplicidade
        solicitacao.anexos.all().delete()
        
        for index, arquivo in enumerate(pdf_termos):
            ext = os.path.splitext(arquivo.name)[1]
            arquivo.name = f"TTI_{safe_termo}_{index+1}{ext}"
            TelefoneSolicitacaoAnexo.objects.create(
                solicitacao=solicitacao,
                arquivo=arquivo,
                ordem=index
            )
            
        # Manter compatibilidade com o serializer salvando apenas que a ação foi registrada no log.
        return Response({'status': 'Solicitação administrativa concluída com sucesso.'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['delete'], url_path='excluir_anexo/(?P<anexo_id>[^/.]+)')
    def excluir_anexo(self, request, pk=None, anexo_id=None):
        solicitacao = self.get_object()
        
        if anexo_id == 'legacy':
            if solicitacao.pdf_termo:
                solicitacao.pdf_termo.delete(save=False)
                solicitacao.pdf_termo = None
            if solicitacao.midia:
                solicitacao.midia.delete(save=False)
                solicitacao.midia = None
            solicitacao.save()
            return Response({'status': 'Anexo legado excluído com sucesso.'}, status=status.HTTP_200_OK)
            
        try:
            anexo = TelefoneSolicitacaoAnexo.objects.get(id=anexo_id, solicitacao=solicitacao)
            anexo.arquivo.delete(save=False)
            anexo.delete()
            return Response({'status': 'Anexo excluído com sucesso.'}, status=status.HTTP_200_OK)
        except TelefoneSolicitacaoAnexo.DoesNotExist:
            return Response({'error': 'Anexo não encontrado.'}, status=status.HTTP_404_NOT_FOUND)


class EmprestimoEventoViewSet(viewsets.ModelViewSet):
    queryset = EmprestimoEvento.objects.all().order_by('-created_at')
    serializer_class = EmprestimoEventoSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        aparelhos_instances = serializer.validated_data.pop('aparelhos', [])
        
        # Validar data_inicio e data_fim
        if 'data_inicio' in request.data:
            dt = parse_datetime(request.data['data_inicio'])
            if dt and timezone.is_naive(dt): dt = timezone.make_aware(dt)
            serializer.validated_data['data_inicio'] = dt
        
        if 'data_fim' in request.data:
            dt = parse_datetime(request.data['data_fim'])
            if dt and timezone.is_naive(dt): dt = timezone.make_aware(dt)
            serializer.validated_data['data_fim'] = dt

        evento = serializer.save()

        # Adicionar os aparelhos e marcar como 'instalado'
        if aparelhos_instances:
            for ap in aparelhos_instances:
                ap.status = 'instalado'
                ap.save()
            evento.aparelhos.set(aparelhos_instances)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['patch'])
    def recolher(self, request, pk=None):
        evento = self.get_object()
        
        if evento.status == 'concluido':
            return Response({'error': 'Evento já foi concluído.'}, status=status.HTTP_400_BAD_REQUEST)

        evento.observacoes = request.data.get('observacoes', evento.observacoes)
        evento.status = 'concluido'
        evento.save()

        # Devolver aparelhos ao estoque
        for ap in evento.aparelhos.all():
            ap.status = 'estoque'
            ap.save()

        return Response({'status': 'Equipamentos recolhidos e evento concluído com sucesso.'}, status=status.HTTP_200_OK)



class RemessaManutencaoViewSet(viewsets.ModelViewSet):
    queryset = RemessaManutencao.objects.all().order_by('-data_remessa')
    serializer_class = RemessaManutencaoSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        aparelhos_ids = serializer.validated_data.pop('aparelhos', [])
        remessa = serializer.save()
        
        for ap_id in aparelhos_ids:
            try:
                ap_voip = AparelhoVoip.objects.get(id=ap_id)
                AparelhoManutencao.objects.create(
                    remessa=remessa,
                    patrimonio=ap_voip.patrimonio,
                    modelo=ap_voip.modelo,
                    fcn=ap_voip.fcn,
                    mac_address=ap_voip.mac_address,
                    ramal=ap_voip.ramal,
                    sala=ap_voip.sala,
                    integridade=ap_voip.integridade,
                    status='manutencao'
                )
                ap_voip.delete()
            except AparelhoVoip.DoesNotExist:
                continue
                
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class CriarSenhaViewSet(viewsets.ModelViewSet):
    queryset = CriarSenha.objects.all().order_by('-created_at')
    serializer_class = CriarSenhaSerializer

    def create(self, request, *args, **kwargs):
        import logging
        logger = logging.getLogger(__name__)
        data = request.data
        colaboradores = data.get('colaboradores')

        if colaboradores and isinstance(colaboradores, list):
            senhas_criadas = []
            for colab in colaboradores:
                # Usa dict() para garantir compatibilidade com QueryDict e JSON
                colab_data = dict(data)
                colab_data.pop('colaboradores', None)  # Remove a lista para nao atrapalhar o serializer
                colab_data.update(colab)
                serializer = self.get_serializer(data=colab_data)
                if not serializer.is_valid():
                    logger.error(f"[CriarSenha.create] Erros de validação: {serializer.errors}")
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                self.perform_create(serializer)
                senhas_criadas.append(serializer.data)
            return Response(senhas_criadas, status=status.HTTP_201_CREATED)
        else:
            return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        senha = serializer.save(status='recebida')

        # Se cargo for colaborador e houver numero_contrato, salva/atualiza o contrato na tabela
        if senha.cargo == 'colaborador' and senha.numero_contrato:
            try:
                contrato, criado = ContratoColaborador.objects.get_or_create(
                    numero_contrato=senha.numero_contrato,
                    defaults={
                        'empresa_vinculada': senha.empresa_vinculada or '',
                        'fiscal_contrato': senha.fiscal_contrato or '',
                        'unidade_fiscal': senha.unidade_fiscal or '',
                    }
                )
            except Exception as e:
                import logging
                logging.getLogger(__name__).warning(f"[perform_create] Erro ao salvar ContratoColaborador: {e}")

        template = TemplateMessage.objects.filter(tipo_evento='tel_solicitacao_senha', is_ativo=True).first()
        if template:
            from notificacoes.models.contato_notificacao import Contato
            from notificacoes.services import disparar_notificacao_contato
            
            contatos = Contato.objects.filter(is_ativo=True, notifica_telefonia=True)
            enviados = set()
            for contato in contatos:
                # Verificação de duplicidade por telefone ou e-mail
                chave_duplicidade = getattr(contato, '_telefone_sanitizado', contato.telefone) or (contato.pessoa.email if contato.pessoa else None)
                if chave_duplicidade:
                    if chave_duplicidade in enviados:
                        continue
                    enviados.add(chave_duplicidade)
                
                texto = template.base_text
                try:
                    text = texto.format(
                        tecnico=contato.nome,
                        protocolo=senha.protocolo or 'N/A',
                        unidade=senha.unidade or 'N/A',
                        sigla_unidade=senha.sigla_unidade or 'N/A',
                        ramal=senha.ramal or 'N/A',
                        usuario=senha.usuario or 'N/A',
                        solicitante=senha.solicitante or 'N/A',
                    )
                    assunto = f"Nova Solicitação de Senha - {senha.protocolo or 'N/A'}"
                    disparar_notificacao_contato(contato, text, text, assunto)
                except Exception as e:
                    print(f"Erro ao formatar/enviar mensagem: {e}")

    def perform_update(self, serializer):
        instance = serializer.save()
        if instance.status == 'aguardando_supervisor' and not instance.nome_tecnico:
            instance.nome_tecnico = self.request.user.get_full_name() or self.request.user.username
            instance.save(update_fields=['nome_tecnico'])

    @action(detail=True, methods=['post'])
    def finalizar(self, request, pk=None):
        senha_obj = self.get_object()
        data = request.data
        
        # Atualiza dados
        senha_obj.cargo = data.get('cargo', senha_obj.cargo)
        senha_obj.numero_contrato = data.get('numero_contrato', senha_obj.numero_contrato)
        senha_obj.empresa_vinculada = data.get('empresa_vinculada', senha_obj.empresa_vinculada)
        senha_obj.fiscal_contrato = data.get('fiscal_contrato', senha_obj.fiscal_contrato)
        senha_obj.unidade_fiscal = data.get('unidade_fiscal', senha_obj.unidade_fiscal)
        senha_obj.status = 'finalizada'
        senha_obj.save()
        
        # Envio de e-mail customizado pelo supervisor
        assunto = data.get('assunto')
        corpo = data.get('corpo')
        to_email = data.get('to_email')
        bcc_email = data.get('bcc_email')
        
        from telefonia.views import enviar_email_senha_manual
        # Mock request to pass data to the old function or just update the function
        request._request.POST = request._request.POST.copy()
        if assunto: request._request.POST['assunto'] = assunto
        if corpo: request._request.POST['corpo'] = corpo
        if to_email: request._request.POST['to_email'] = to_email
        if bcc_email: request._request.POST['bcc_email'] = bcc_email
        
        try:
            resp = enviar_email_senha_manual(request._request, pk=senha_obj.pk)
            import json
            resp_content = json.loads(resp.content)
            if resp.status_code == 200:
                return Response({'status': 'Senha finalizada e e-mail enviado.'}, status=status.HTTP_200_OK)
            else:
                return Response({'error': resp_content.get('error', 'Erro ao enviar email.')}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'Erro ao finalizar: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='email-preview')
    def email_preview(self, request, pk=None):
        senha_obj = self.get_object()
        from telefonia.models.padrao_email import PadraoEmailTelefonia
        padrao = PadraoEmailTelefonia.objects.filter(ativo=True).first()
        if not padrao:
            return Response({"error": "Nenhum template de e-mail ativo."}, status=400)
            
        try:
            corpo_formatado = padrao.corpo.format(
                primeiro_nome=senha_obj.primeiro_nome or 'Usuário',
                senha=senha_obj.senha or 'Não registrada'
            )
        except:
            corpo_formatado = padrao.corpo
            
        corpo = f"{corpo_formatado}\n\n{padrao.assinatura}" if padrao.assinatura else corpo_formatado
        
        from django.utils.text import slugify
        nome_usuario = slugify(senha_obj.usuario) if senha_obj.usuario and senha_obj.usuario != "N/A" else "documento_senha"
        
        return Response({
            'to_email': senha_obj.email or '',
            'bcc_email': padrao.email_copia or '',
            'assunto': padrao.assunto or '',
            'corpo': corpo,
            'filename': f"{nome_usuario}.pdf"
        })

class ContratoColaboradorViewSet(viewsets.ModelViewSet):
    queryset = ContratoColaborador.objects.all().order_by('-created_at')
    serializer_class = ContratoColaboradorSerializer
    pagination_class = None


# PDF GENERATION
def gerar_pdf_remessa(request, pk):
    try:
        remessa = RemessaManutencao.objects.get(pk=pk)
    except RemessaManutencao.DoesNotExist:
        return HttpResponse("Remessa não encontrada", status=404)
        
    context = {'remessa': remessa}
    template = get_template('telefonia/pdf_remessa.html')
    html_string = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="remessa_{remessa.id}.pdf"'
        return response
    return HttpResponse("Erro ao gerar PDF", status=500)


def get_pdf_senha_bytes(senha_id=None, template_id=None):
    senha = CriarSenha.objects.get(pk=senha_id) if senha_id else None
    padrao = PadraoSenhaTelefonia.objects.get(pk=template_id) if template_id else PadraoSenhaTelefonia.objects.filter(ativo=True).first()
    
    if not padrao:
        return None
        
    context = {
        'senha': senha,
        'padrao': padrao,
        'data_atual': datetime.date.today().strftime('%d.%m.%Y')
    }
    
    template = get_template('telefonia/pdf_senha.html')
    html_string = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
    
    if not pdf.err:
        return result.getvalue()
    return None


def gerar_pdf_senha(request, pk):
    pdf_bytes = get_pdf_senha_bytes(senha_id=pk)
    if pdf_bytes:
        senha = CriarSenha.objects.get(pk=pk)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        from django.utils.text import slugify
        nome_usuario = slugify(senha.usuario) if senha.usuario and senha.usuario != "N/A" else "usuario"
        response['Content-Disposition'] = f'inline; filename="{nome_usuario}.pdf"'
        return response
    return HttpResponse("Erro ao gerar PDF", status=500)


def preview_pdf_senha_template(request, template_id):
    pdf_bytes = get_pdf_senha_bytes(template_id=template_id)
    if pdf_bytes:
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="preview_senha.pdf"'
        return response
    return HttpResponse("Erro ao gerar preview", status=500)


def get_pdf_tutorial_bytes(template_id=None):
    padrao = PadraoTutorialTelefonia.objects.get(pk=template_id) if template_id else PadraoTutorialTelefonia.objects.filter(ativo=True).first()
    if not padrao:
        return None

    context = {'padrao': padrao}
    template = get_template('telefonia/pdf_tutorial.html')
    html_string = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
    
    if not pdf.err:
        return result.getvalue()
    return None


def gerar_pdf_tutorial(request):
    pdf_bytes = get_pdf_tutorial_bytes()
    if pdf_bytes:
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="Tutorial - Validação de Ligações A Serviço e Particulares.pdf"'
        return response
    return HttpResponse("Erro ao gerar PDF", status=500)


def preview_pdf_tutorial_template(request, template_id):
    pdf_bytes = get_pdf_tutorial_bytes(template_id=template_id)
    if pdf_bytes:
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="preview_tutorial.pdf"'
        return response
    return HttpResponse("Erro ao gerar preview", status=500)


def enviar_email_senha_manual(request, pk):
    from django.core.mail import EmailMessage
    from django.conf import settings
    try:
        senha_obj = CriarSenha.objects.get(pk=pk)
    except CriarSenha.DoesNotExist:
        return JsonResponse({"error": "Senha não encontrada."}, status=404)
        
    if not senha_obj.email:
        return JsonResponse({"error": "O usuário não possui um e-mail cadastrado."}, status=400)
        
    padrao_email = PadraoEmailTelefonia.objects.filter(ativo=True).first()
    if not padrao_email:
        return JsonResponse({"error": "Nenhum template de e-mail ativo foi encontrado."}, status=400)
        
    pdf_senha_bytes = get_pdf_senha_bytes(senha_id=pk)
    pdf_tutorial_bytes = get_pdf_tutorial_bytes()
    
    if not pdf_senha_bytes or not pdf_tutorial_bytes:
        return JsonResponse({"error": "Erro ao gerar os documentos em PDF para anexo."}, status=500)
        
    try:
        corpo_formatado = padrao_email.corpo.format(
            primeiro_nome=senha_obj.primeiro_nome or 'Usuário',
            senha=senha_obj.senha or 'Não registrada'
        )
    except Exception:
        corpo_formatado = padrao_email.corpo
        
    corpo_default = f"{corpo_formatado}\n\n{padrao_email.assinatura}" if padrao_email.assinatura else corpo_formatado
        
    assunto = request.POST.get('assunto') or padrao_email.assunto
    corpo = request.POST.get('corpo') or corpo_default
    
    to_email_val = request.POST.get('to_email') or senha_obj.email
    to_email = [to_email_val] if to_email_val else []
    
    bcc_email_val = request.POST.get('bcc_email') or padrao_email.email_copia
    bcc_email = [bcc_email_val] if bcc_email_val else []
    
    email = EmailMessage(
        subject=assunto,
        body=corpo,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=to_email,
        bcc=bcc_email
    )
    
    nome_usuario = senha_obj.usuario if senha_obj.usuario and senha_obj.usuario != "N/A" else "Usuario"
    email.attach(f'{nome_usuario}.pdf', pdf_senha_bytes, 'application/pdf')
    email.attach('Tutorial - Validação de Ligações A Serviço e Particulares.pdf', pdf_tutorial_bytes, 'application/pdf')
    
    try:
        email.send(fail_silently=False)
        return JsonResponse({"message": "E-mail enviado com sucesso!"})
    except Exception as e:
        return JsonResponse({"error": f"Erro ao enviar o e-mail: {str(e)}"}, status=500)
import io
import base64
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from django.template.loader import get_template
from django.views import View
from django.utils import timezone
from xhtml2pdf import pisa

from telefonia.models import AparelhoVoip

class AparelhosReportView(View):
    def get(self, request, *args, **kwargs):
        formato = request.GET.get('formato', 'pdf').lower()
        
        # Pega todos os aparelhos
        aparelhos = AparelhoVoip.objects.all()
        
        # Estatísticas para o gráfico
        counts = {
            'estoque': aparelhos.filter(status='estoque').count(),
            'instalado': aparelhos.filter(status='instalado').count(),
            'manutencao': aparelhos.filter(status='manutencao').count(),
            'defeituoso': aparelhos.filter(status='defeituoso').count(),
        }
        
        # Filtra a lista principal para exibir apenas os instalados
        aparelhos_lista = aparelhos.filter(status='instalado').order_by('sala', 'patrimonio')
        
        if formato == 'excel':
            return self.export_excel(counts, aparelhos_lista)
        else:
            return self.export_pdf(counts, aparelhos_lista)

    def generate_chart_base64(self, counts):
        labels = ['Em Estoque', 'Instalados', 'Em Manutenção', 'Defeituosos']
        sizes = [counts['estoque'], counts['instalado'], counts['manutencao'], counts['defeituoso']]
        colors = ['#007bff', '#28a745', '#ffc107', '#dc3545']
        
        # Remove zeros to prevent empty slices
        filtered_labels = [l for l, s in zip(labels, sizes) if s > 0]
        filtered_sizes = [s for s in sizes if s > 0]
        filtered_colors = [c for c, s in zip(colors, sizes) if s > 0]
        
        fig, ax = plt.subplots(figsize=(6, 4))
        if filtered_sizes:
            ax.pie(filtered_sizes, labels=filtered_labels, colors=filtered_colors, autopct='%1.1f%%', startangle=140)
            ax.axis('equal')
        else:
            ax.text(0.5, 0.5, 'Nenhum dado disponível', horizontalalignment='center', verticalalignment='center')
        
        plt.title('Desempenho da Gestão de Aparelhos')
        
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', transparent=True, bbox_inches='tight')
        plt.close(fig)
        buffer.seek(0)
        
        img_str = base64.b64encode(buffer.read()).decode('utf-8')
        return f"data:image/png;base64,{img_str}"

    def export_pdf(self, counts, aparelhos_lista):
        template = get_template('telefonia/relatorios/relatorio_aparelhos_pdf.html')
        chart_uri = self.generate_chart_base64(counts)
        
        context = {
            'counts': counts,
            'aparelhos': aparelhos_lista,
            'chart_uri': chart_uri,
            'data_emissao': timezone.now()
        }
        
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="relatorio_aparelhos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        pisa_status = pisa.CreatePDF(
            html, dest=response
        )
        if pisa_status.err:
            return HttpResponse('Tivemos alguns erros <pre>' + html + '</pre>')
        return response

    def export_excel(self, counts, aparelhos_lista):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aparelhos Instalados"
        
        # Header da tabela
        headers = ['Patrimônio', 'Modelo', 'MAC Address', 'Ramal', 'Integridade', 'Local/Sala']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Popula a tabela
        row_num = 2
        for aparelho in aparelhos_lista:
            ws.cell(row=row_num, column=1, value=aparelho.patrimonio or '-')
            ws.cell(row=row_num, column=2, value=aparelho.modelo or '-')
            ws.cell(row=row_num, column=3, value=aparelho.mac_address or '-')
            ws.cell(row=row_num, column=4, value=aparelho.ramal or '-')
            ws.cell(row=row_num, column=5, value=aparelho.get_integridade_display() if aparelho.integridade else '-')
            ws.cell(row=row_num, column=6, value=aparelho.sala or '-')
            row_num += 1
            
        # Aba de Gráfico
        ws_chart = wb.create_sheet(title="Desempenho (Gráfico)")
        ws_chart.cell(row=1, column=1, value="Status")
        ws_chart.cell(row=1, column=2, value="Quantidade")
        
        status_data = [
            ("Em Estoque", counts['estoque']),
            ("Instalados", counts['instalado']),
            ("Em Manutenção", counts['manutencao']),
            ("Defeituosos", counts['defeituoso']),
        ]
        
        for idx, (label, count) in enumerate(status_data, 2):
            ws_chart.cell(row=idx, column=1, value=label)
            ws_chart.cell(row=idx, column=2, value=count)
            
        pie = PieChart()
        labels = Reference(ws_chart, min_col=1, min_row=2, max_row=5)
        data = Reference(ws_chart, min_col=2, min_row=1, max_row=5)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Desempenho da Gestão de Aparelhos"
        
        ws_chart.add_chart(pie, "D2")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=relatorio_aparelhos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        
        return response
